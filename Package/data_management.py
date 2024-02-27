from collections import Counter
import os

import datetime
import openpyxl as op
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import customtkinter as ctk
import darkdetect
from .styles import *
from .note_management import NotesManager
from .achievement import Achievement

class DataManager:
    def __init__(self, App, timer_manager, workbook, worksheet):
        self.app = App
        self.timer_manager = timer_manager
        self.workbook = workbook
        self.worksheet = worksheet
        self.initialize_variables()

        self.notes_manager = NotesManager(self.app, self)


    def initialize_variables(self) -> None:
        self.day_name_list = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        self.date_list = []
        self.duration_list = []
        self.total_duration = 0
        self.total_break_duration = 0
        self.best_weekday = ""
        self.average_time = "00:00"
        self.graph_color = "#f38064"
        self.graph_bg_color = graph_bg_color
        self.graph_fg_color = graph_fg_color
        self.font_color = font_color
        self.spine_color = border_frame_color
        

    def initialize_new_file_variables(self) -> None:
        self.goal_amount = 0
        self.data_amount = 0
        self.notes_amount = 0
        self.monday_duration = self.tuesday_duration = self.wednesday_duration = self.thursday_duration = self.friday_duration = self.saturday_duration = self.sunday_duration = 0
        self.achievements = []
        self.color_name = "Orange"
        if darkdetect.isDark():
            self.theme_name = "Dark"
        else:
            self.theme_name = "Light"
        self.customize_excel()
        self.save_color()
        self.save_theme()

    
    def collect_data(self) -> None:
        self.data_amount = int(self.worksheet["Z2"].value)
        self.goal_amount = int(self.worksheet["R2"].value)
        self.notes_amount = int(self.worksheet["N9"].value)

        self.collect_day_data()

        print("Data collected.")


    def data_to_variable(self) -> None:
        self.clear_graph_lists()

        for data in range(2, self.data_amount + 2):
            if "/" in str(self.worksheet["B" + str(data)].value):
                self.date_list.append(datetime.datetime.strptime(str(self.worksheet["B" + str(data)].value).split(" ")[0], "%d/%m/%Y").date())
            elif "-" in str(self.worksheet["B" + str(data)].value):
                self.date_list.append(datetime.datetime.strptime(str(self.worksheet["B" + str(data)].value).split(" ")[0], "%Y-%m-%d").date())
            self.duration_list.append(round(self.worksheet["C" + str(data)].value))

        self.total_duration = sum(self.duration_list)
        
        self.create_total_data()


    def create_total_data(self):
        self.break_list = []
        self.hours_list = []
        self.subject_list = []
        def get_sec(time: str) -> int:
            """Get seconds from time."""
            h, m = time.split(":")
            return int(h) * 3600 + int(m) * 60
        
        for data in range(2, self.data_amount + 2):
            self.break_list.append(float(self.worksheet["D" + str(data)].value))
            self.hours_list.append(get_sec((self.worksheet["A" + str(data)].value.split(" ")[1])))
            self.subject_list.append(self.worksheet["E" + str(data)].value)
        try:           
            self.total_break_duration = round(sum(self.break_list))
        except:
            self.total_break_duration = 0
        try:
            self.average_time = str(datetime.timedelta(seconds=(round(sum(self.hours_list) / len(self.hours_list)))))[:5]
        except ZeroDivisionError:
            self.average_time = "00:00"
        try:
            self.most_common_subject = Counter(self.subject_list).most_common(1)[0][0]
        except IndexError:
            self.most_common_subject = ""
        try:
            self.most_common_subject_amount = Counter(self.subject_list).most_common(1)[0][1]
        except:
            self.most_common_subject_amount = 0
        try:
            self.unique_subjects = set(self.subject_list)
        except:
            self.unique_subjects = []
        try:
            self.longest_session = round(max(self.duration_list))
        except:
            self.longest_session = 0
        

        self.create_achievements()
        

        def get_weekday():
            weekdays_dict = {}
            for i in range(2, 9):
                if self.worksheet["W" + str(i)].value != 0:
                    weekdays_dict[self.worksheet["W" + str(i)].value] = self.day_name_list[i-2]
            if weekdays_dict:
                self.best_weekday = weekdays_dict[max(weekdays_dict)]
            else:
                self.best_weekday = ""
        get_weekday()



    def save_data(self) -> None:
        self.initialize_variables()
        self.set_weekday()
        self.save_color()
        
        self.data_amount += 1

        self.stop_time = datetime.datetime.now()

        self.duration = self.calculate_duration()

        self.workbook.save(self.app.data_file)

        self.write_to_excel()

        print("Data saved.")

        self.timer_manager.initialize_variables()
        self.app.reset_timers()


    def write_to_excel(self) -> None:
        self.worksheet["A" + str((self.data_amount + 1))].value = self.start_time.strftime("%d/%m/%Y %H:%M")
        self.worksheet["B" + str((self.data_amount + 1))].value = self.stop_time.strftime("%d/%m/%Y %H:%M")
        self.worksheet["C" + str((self.data_amount + 1))].value = self.duration
        self.worksheet["D" + str((self.data_amount + 1))].value = self.timer_manager.break_time/60
        self.worksheet["E" + str((self.data_amount + 1))].value = self.app.subject_selection.get()

        self.worksheet["N9"].value = self.notes_amount

        self.worksheet["R2"].value = self.goal_amount

        self.worksheet["Z2"].value = self.data_amount
        self.workbook.save(self.app.data_file)


    def customize_excel(self) -> None:
        self.worksheet["A1"].value = "Start:"
        self.worksheet["B1"].value = "End:"
        self.worksheet["C1"].value = "Duration:"
        self.worksheet["D1"].value = "Break:"
        self.worksheet["E1"].value = "Subject:"

        self.worksheet["N8"].value = "Notes amount:"
        self.worksheet["N9"].value = self.notes_amount
        self.worksheet["N11"].value = "Notes:"
        self.worksheet["M12"].value = "Deleted:"
        self.worksheet["N12"].value = "Date:"
        self.worksheet["O12"].value = "Title:"
        self.worksheet["P12"].value = "Text:"

        self.worksheet["P1"].value = "Autobreak:"

        self.worksheet["Q1"].value = "Eye care:"
        self.worksheet["Q4"].value = "Only when timer running:"

        self.worksheet["R1"].value = "Goals reached:"
        self.worksheet["R2"].value = self.goal_amount

        self.worksheet["S1"].value = "Subject:"

        self.worksheet["T1"].value = "Color:"
        self.worksheet["T2"].value = self.color_name

        self.worksheet["U1"].value = "Theme:"
        self.worksheet["U2"].value = self.theme_name

        self.worksheet["W1"].value = "Weekday duration:"

        self.worksheet["Z1"].value = "Data amount: "
        self.worksheet["Z1"].font = Font(bold=True, size=14)
        self.worksheet["Z2"].value = self.data_amount

        self.save_weekday_data()
        self.style_excel()

        self.workbook.save(self.app.data_file)
        print("Excel customized.")


    def style_excel(self) -> None:
        self.worksheet["A1"].font = Font(bold=True, size=14)
        self.worksheet["B1"].font = Font(bold=True, size=14)
        self.worksheet["C1"].font = Font(bold=True, size=14)
        self.worksheet["D1"].font = Font(bold=True, size=14)
        self.worksheet["E1"].font = Font(bold=True, size=14)


    def get_start_time(self) -> None:
        self.start_time = datetime.datetime.now()


    def calculate_duration(self) -> float:
        duration = self.timer_manager.timer_time
        if duration < 0: 
            duration = 0
        else:
            duration /= 60
        return duration
    

    def increase_goal_streak(self) -> None:
        self.goal_amount += 1


    def clear_graph_lists(self) -> None:
        self.date_list.clear()
        self.duration_list.clear()


    def collect_day_data(self) -> None:
        self.monday_duration = round(self.worksheet["W2"].value)
        self.tuesday_duration = round(self.worksheet["W3"].value)
        self.wednesday_duration = round(self.worksheet["W4"].value)
        self.thursday_duration = round(self.worksheet["W5"].value)
        self.friday_duration = round(self.worksheet["W6"].value)
        self.saturday_duration = round(self.worksheet["W7"].value)
        self.sunday_duration = round(self.worksheet["W8"].value)

        self.day_duration_list = [self.monday_duration, self.tuesday_duration, self.wednesday_duration, self.thursday_duration, self.friday_duration, self.saturday_duration, self.sunday_duration]
    

    def set_weekday(self) -> None:
        duration = self.calculate_duration()
        weekday_today = datetime.datetime.now().weekday()

        match weekday_today:
            case 0:
                self.monday_duration += duration
            case 1:
                self.tuesday_duration += duration
            case 2:
                self.wednesday_duration += duration
            case 3:
                self.thursday_duration += duration
            case 4:
                self.friday_duration += duration
            case 5:
                self.saturday_duration += duration
            case 6:
                self.sunday_duration += duration

        self.save_weekday_data()

        self.workbook.save(self.app.data_file)
        print("Weekday saved.")


    def save_weekday_data(self) -> None:
        self.worksheet["W2"].value = self.monday_duration
        self.worksheet["W3"].value = self.tuesday_duration
        self.worksheet["W4"].value = self.wednesday_duration
        self.worksheet["W5"].value = self.thursday_duration
        self.worksheet["W6"].value = self.friday_duration
        self.worksheet["W7"].value = self.saturday_duration
        self.worksheet["W8"].value = self.sunday_duration


    def save_autobreak(self, frequency: str, duration: str, switch: str) -> None:
        self.worksheet["P2"].value = switch
        self.worksheet["P3"].value = int(frequency)
        self.worksheet["P4"].value = int(duration)

        self.workbook.save(self.app.data_file)

        print("Autobreak saved.")

        self.load_autobreak()
        

    def load_autobreak(self):
        self.autobreak_on = self.worksheet["P2"].value
        self.autobreak_frequency = self.worksheet["P3"].value
        self.autobreak_duration = self.worksheet["P4"].value

        if self.worksheet["P2"].value == None:
            self.autobreak_on = "Off"

        if self.autobreak_frequency == None:
            self.autobreak_frequency = 25

        if self.autobreak_duration == None:
            self.autobreak_duration = 5

        self.app.frequency_input.configure(placeholder_text=self.autobreak_frequency)
        self.app.duration_input.configure(placeholder_text=self.autobreak_duration)
        self.app.autobreak_switch.configure(variable=ctk.StringVar(value=self.autobreak_on))

        self.app.WINDOW.after(0, self.app.auto_break)

        print("Autobreak loaded.")


    def set_color(self, color_dropdown) -> None:
        self.color_name = color_dropdown.get()
        if self.color_name != self.worksheet["T2"].value:
            print("Color set.")
            self.save_color()


    def save_color(self) -> None:
        self.worksheet["T2"].value = self.color_name
        self.load_color()


    def load_color(self) -> None:
        self.color_name = self.worksheet["T2"].value
        self.app.color_dropdown.configure(variable=ctk.StringVar(value=self.color_name))
        colors = {"Orange": [orange_button_color, orange_highlight_color, orange_pie_colors], 
                    "Green": [green_button_color, green_highlight_color, green_pie_colors], 
                    "Blue": [blue_button_color, blue_highlight_color, blue_pie_colors],
                    "Pink": [pink_button_color, pink_highlight_color, pink_pie_colors]}
        
        self.color = colors[self.color_name][0]
        self.highlight_color = colors[self.color_name][1]
        self.pie_colors = colors[self.color_name][2]
        self.graph_color = self.color
        print("Color loaded.")
        self.change_color()


    def change_color(self) -> None:
        for widget in self.app.widget_list:
            if isinstance(widget, ctk.CTkButton):
                widget.configure(fg_color=self.color, hover_color=self.highlight_color)
            else:
                if widget.get() != 0:
                    widget.configure(progress_color = self.color)
                else:
                    widget.configure(progress_color=(light_border_frame_color, border_frame_color))
        #self.app.progressbar.configure(progress_color = self.color)
        self.app.eye_care_checkbox.configure(fg_color=self.color)
        self.app.create_graphs()

        self.app.create_achievements()

        self.load_notes()
        print("Color changed.")


    def set_theme(self, theme_dropdown) -> None:
        self.theme_name = theme_dropdown.get()
        if self.theme_name != self.worksheet["U2"].value:
            print("Theme set.")
            self.save_theme()


    def save_theme(self) -> None:
        self.worksheet["U2"].value = self.theme_name
        self.load_theme()

    
    def load_theme(self) -> None:
        self.theme_name = self.worksheet["U2"].value
        self.app.theme_dropdown.configure(variable=ctk.StringVar(value=self.theme_name))

        if self.theme_name == "Dark":
            ctk.set_appearance_mode("dark")
            self.graph_bg_color = graph_bg_color
            self.graph_fg_color = graph_fg_color
            self.spine_color = border_frame_color
            self.font_color = "white"
        else:
            ctk.set_appearance_mode("light")
            self.graph_bg_color = light_graph_bg_color
            self.graph_fg_color = light_graph_fg_color
            self.spine_color = light_border_frame_color
            self.font_color = "black"

        self.app.create_graphs()

        print("Theme loaded.")


    def save_subject(self, subject: str) -> None:
        self.worksheet["S2"].value = subject
        print("Subject saved.")


    def load_subject(self) -> None:
        if self.worksheet["S2"].value != None:
            subject = self.worksheet["S2"].value
        else:
            subject = "Other"

        self.app.subject_selection.configure(variable=ctk.StringVar(value=subject))

        return subject
    

    def create_new_note(self, title, text):
        self.notes_amount += 1
        
        self.worksheet["N9"].value = self.notes_amount

        self.worksheet["N" + str(self.notes_amount + 12)].value = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        self.worksheet["O" + str(self.notes_amount + 12)].value = title
        self.worksheet["P" + str(self.notes_amount + 12)].value = text

        self.workbook.save(self.app.data_file)

        print("New note created.")

        self.create_achievements()
        self.app.create_achievements()
        self.load_notes()


    def load_notes(self) -> None:
        self.notes_manager.clear_notes()

        if self.notes_amount == 0:
            return None
        
        for i in range(self.notes_amount+12, 12, -1):
            if self.worksheet["M" + str(i)].value != "Yes":
                self.notes_manager.create_task(i)


    def save_eye_care(self, eye_care: str, checkbox: str) -> None:
        self.worksheet["Q2"].value = eye_care
        self.worksheet["Q5"].value = checkbox

        self.workbook.save(self.app.data_file)
        print("Eye care saved.")


    def load_eye_care(self) -> None:
        if self.worksheet["Q2"].value != None:
            eye_care = self.worksheet["Q2"].value
        else:
            eye_care = "Off"

        self.app.eye_care_selection.configure(variable=ctk.StringVar(value=eye_care))

        if self.worksheet["Q5"].value != None:
            checkbox = self.worksheet["Q5"].value
        else:
            checkbox = "Off"

        self.app.eye_care_checkbox.configure(variable=ctk.StringVar(value=checkbox))

        self.app.WINDOW.after(60*20*1000, self.app.eye_protection)  # Schedule initial iteration for eye_protection


    def export_data(self):
        def change_cell_width(worksheet, cell_range: tuple, width: int = 15) -> None:
            start, end = cell_range
            for cell in range(start, end + 1):
                cell_letter = get_column_letter(cell)
                worksheet.column_dimensions[cell_letter].width = width

        def align_cells(worksheet, cell_range: str):
            for row in worksheet[cell_range]:
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
        export_workbook = op.Workbook()
        timer_worksheet = export_workbook.active
        timer_worksheet.title = "Timer"
        notes_worksheet = export_workbook.create_sheet("Notes")

        timer_worksheet.merge_cells("A1:E1")
        timer_worksheet["A1"] = "Timer"
        timer_worksheet["A1"].font = Font(bold=True, size=16)
        align_cells(timer_worksheet, "A1:E1")

        timer_worksheet["A2"].value = "Start:"
        timer_worksheet["A2"].font = Font(bold=True, size=14)
        timer_worksheet["B2"].value = "End:"
        timer_worksheet["B2"].font = Font(bold=True, size=14)
        timer_worksheet["C2"].value = "Duration:"
        timer_worksheet["C2"].font = Font(bold=True, size=14)
        timer_worksheet["D2"].value = "Break Duration:"
        timer_worksheet["D2"].font = Font(bold=True, size=14)
        timer_worksheet["E2"].value = "Subject:"
        timer_worksheet["E2"].font = Font(bold=True, size=14)
        change_cell_width(timer_worksheet, (1, 5), 20)

        notes_worksheet.merge_cells("A1:C1")
        notes_worksheet["A1"] = "Notes"
        notes_worksheet["A1"].font = Font(bold=True, size=16)
        align_cells(notes_worksheet, "A1:C1")

        notes_worksheet["A2"].value = "Date:"
        notes_worksheet["A2"].font = Font(bold=True, size=14)
        notes_worksheet["B2"].value = "Title:"
        notes_worksheet["B2"].font = Font(bold=True, size=14)
        notes_worksheet["C2"].value = "Text:"
        notes_worksheet["C2"].font = Font(bold=True, size=14)
        change_cell_width(notes_worksheet, (1, 3), 20)

        for data in range(3, self.data_amount + 3):
            timer_worksheet["A" + str(data)].value = self.worksheet["A" + str(data - 1)].value
            timer_worksheet["B" + str(data)].value = self.worksheet["B" + str(data - 1)].value
            timer_worksheet["C" + str(data)].value = str(round(self.worksheet["C" + str(data - 1)].value, 1)) + "m"
            timer_worksheet["D" + str(data)].value = str(round(self.worksheet["D" + str(data - 1)].value, 1)) + "m"
            timer_worksheet["E" + str(data)].value = self.worksheet["E" + str(data - 1)].value

        note_list = []
        for note in range(self.notes_amount + 12, 12, -1):
            if self.worksheet["M" + str(note)].value == "Yes":
                continue
            else:
                note_list.append(note)


        for index, note in enumerate(note_list):
            notes_worksheet["A" + str(index + 3)].value = self.worksheet["N" + str(note)].value
            notes_worksheet["A" + str(index + 3)].alignment = Alignment(horizontal="left", vertical="top")
            notes_worksheet["A" + str(index + 3)].font = Font(size=12)
            notes_worksheet["B" + str(index + 3)].value = self.worksheet["O" + str(note)].value
            notes_worksheet["B" + str(index + 3)].alignment = Alignment(horizontal="left", vertical="top")
            notes_worksheet["B" + str(index + 3)].font = Font(size=12)
            notes_worksheet["C" + str(index + 3)].value = self.worksheet["P" + str(note)].value
            notes_worksheet["C" + str(index + 3)].alignment =  Alignment(horizontal="left", vertical="top", wrap_text=True)
            notes_worksheet.column_dimensions["C"].width = 107

        export_workbook.save(f"{os.path.join(os.path.expanduser("~"), "Desktop")}/timer_data_{datetime.datetime.now().date().strftime("%d.%m.%Y")}.xlsx")

        print("File exported.")


    def create_achievements(self):
        self.achievements = [Achievement(name="Time Titan", title="Clock in 1000 minutes of study, mastering the art of time management.", max_value=1000, value=self.total_duration),
                             Achievement(name="Goal Getter", title="Reach 30 goals, proving dedication to progress.", max_value=30, value=self.goal_amount),
                             Achievement(name="Subject Explorer", title="Dive into 7 different subjects, broadening your knowledge horizons.", max_value=7, value=len(self.unique_subjects)),
                             Achievement(name="Focus Maestro", title="Master concentration in a 5-hour session, demonstrating exceptional focus.", max_value=5, value=round(self.longest_session/60, 2)),
                             Achievement(name="Subject Savant", title="Study one subject 30 times, becoming a savant in its intricacies.", max_value=30, value=self.most_common_subject_amount),
                             Achievement(name="Restful Respite", title="Accumulate 200 minutes of break time, rejuvenating your mind and body.", max_value=200, value=self.total_break_duration), 
                             Achievement(name="Daily Discipline", title="Exhibit discipline through diligent study for 30 days.", max_value=30, value=len(set(self.date_list))),
                             Achievement(name="Note Scribbler", title="Scribble down 10 notes, capturing key insights and ideas.", max_value=10, value=self.notes_amount)]