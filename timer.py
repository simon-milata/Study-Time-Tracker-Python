import datetime
import openpyxl as op
from openpyxl.styles import Font
import os
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import customtkinter as ctk
from styles import *
from matplotlib.ticker import MaxNLocator

APPNAME = "Timer App"
FILENAME = "Timer Data.xlsx"

local_folder = os.path.expandvars(rf"%APPDATA%\{APPNAME}")
data_file = os.path.expandvars(rf"%APPDATA%\{APPNAME}\{FILENAME}")

os.makedirs(local_folder, exist_ok=True)


WINDOW = ctk.CTk()
WINDOW.geometry(str(WIDTH + BORDER_WIDTH + main_frame_pad_x + tab_frame_width) + "x" + str(HEIGHT+((widget_padding_x+frame_padding)*2)))
WINDOW.title(APPNAME)
WINDOW.configure(background=window_color)
WINDOW.resizable(False, False)
WINDOW.grid_propagate(False)


#---------------------------------------------------------------------------VARIABLES---------------------------------------------------------------------------#
timer_running = False
break_running = False
timer_time = 0
break_time = 0
start_time = ""
goal = 0
default_choice = ctk.StringVar(value="1 hour")
color = "Orange"
default_color = ctk.StringVar(value=color)


main_frame = ctk.CTkFrame(WINDOW, fg_color=main_frame_color, height=HEIGHT+((widget_padding_x+frame_padding)*2), width=WIDTH, corner_radius=0)
main_frame.grid(column=2, row=0, padx=main_frame_pad_x)
main_frame.grid_propagate(False)

statistics_frame = ctk.CTkFrame(WINDOW, fg_color=main_frame_color, height=HEIGHT+((widget_padding_x+frame_padding)*2), width=WIDTH, corner_radius=0)
statistics_frame.grid(column=2, row=0, padx=main_frame_pad_x)
statistics_frame.grid_forget()
statistics_frame.grid_propagate(False)

settings_frame = ctk.CTkFrame(WINDOW, fg_color=main_frame_color, height=HEIGHT+((widget_padding_x+frame_padding)*2), width=WIDTH, corner_radius=0)
settings_frame.grid(column=2, row=0, padx=main_frame_pad_x)
settings_frame.grid_forget()
settings_frame.grid_propagate(False)

def customize_excel(worksheet):
    worksheet["A1"].value = "Start:"
    worksheet["B1"].value = "End:"
    worksheet["C1"].value = "Duration:"
    worksheet["D1"].value = "Break:"
    worksheet["E1"].value = "Streak:"

    worksheet["U1"].value = "Monday:"
    worksheet["U2"].value = "Tuesday:"
    worksheet["U3"].value = "Wednesday:"
    worksheet["U4"].value = "Thursday:"
    worksheet["U5"].value = "Friday:"
    worksheet["U6"].value = "Saturday:"
    worksheet["U7"].value = "Sunday:"

    worksheet["S1"].value = "Color:"

    worksheet["V1"].value = monday_amount
    worksheet["V2"].value = tuesday_amount
    worksheet["V3"].value = wednesday_amount
    worksheet["V4"].value = thursday_amount
    worksheet["V5"].value = friday_amount
    worksheet["V6"].value = saturday_amount
    worksheet["V7"].value = sunday_amount

    worksheet["W1"].value = monday_duration
    worksheet["W2"].value = tuesday_duration
    worksheet["W3"].value = wednesday_duration
    worksheet["W4"].value = thursday_duration
    worksheet["W5"].value = friday_duration
    worksheet["W6"].value = saturday_duration
    worksheet["W7"].value = sunday_duration

    worksheet["T1"].value = color

    worksheet["R1"].value = goal_amount

    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["B1"].font = Font(bold=True, size=14)
    worksheet["C1"].font = Font(bold=True, size=14)
    worksheet["D1"].font = Font(bold=True, size=14)
    worksheet["E1"].font = Font(bold=True, size=14)

    worksheet["X1"].value = "Data amount: "
    worksheet["X1"].font = Font(bold=True, size=14)
    worksheet["Z1"].value = data_amount
    workbook.save(data_file)
    print("Excel customized.")

date_list = []
duration_list = []
total_duration = 0


def create_time_spent_graph(date_list, duration_list):
    global graph_color
    data = {"Date": date_list, "Duration": duration_list}
    df = pd.DataFrame(data)
    grouped_data = df.groupby("Date")["Duration"].sum().reset_index()
    fig1, ax1 = plt.subplots()
    ax1.bar(grouped_data["Date"], grouped_data["Duration"], color=graph_color)
    ax1.set_xlabel("Date", color=font_color)
    ax1.set_ylabel("Duration in minutes", color=font_color)
    ax1.set_title("Time Spent by Date", color=font_color)
    ax1.tick_params(colors="white")
    ax1.set_facecolor(graph_fg_color)
    fig1.set_facecolor(graph_bg_color)
    ax1.spines["top"].set_color(spine_color)
    ax1.spines["bottom"].set_color(spine_color)
    ax1.spines["left"].set_color(spine_color)
    ax1.spines["right"].set_color(spine_color)
    fig1.set_size_inches(graph_width/100, graph_height/100, forward=True)
    ax1.set_xticklabels(grouped_data["Date"], rotation=45, ha='right')

    date_format = mdates.DateFormatter("%d/%m")
    ax1.xaxis.set_major_formatter(date_format)
    ax1.xaxis.set_major_locator(MaxNLocator(integer=True, prune='both'))
    time_spent_frame = FigureCanvasTkAgg(fig1, master=statistics_frame)
    plt.subplots_adjust(bottom=0.2)

    time_spent_graph = time_spent_frame.get_tk_widget()
    time_spent_graph.grid(row=0, column=0, padx=10, pady=10)
    time_spent_graph.config(highlightbackground=frame_border_color, highlightthickness=2, background=frame_color)

    date_list.clear()
    duration_list.clear()


def calculate_duration():
    global timer_time, break_time
    duration = timer_time - break_time
    if duration < 0: 
        duration = 0
    else:
        duration /= 60
    return duration


def save_weekday():
    global timer_time, break_time
    global monday_amount, tuesday_amount, wednesday_amount, thursday_amount, friday_amount, saturday_amount, sunday_amount
    global monday_duration, tuesday_duration, wednesday_duration, thursday_duration, friday_duration, saturday_duration, sunday_duration

    duration = calculate_duration()
    print("AAAAA")
    print(duration)

    match datetime.datetime.now().weekday():
        case 0:
            print("Case 0")
            monday_amount += 1
            worksheet["V1"].value = monday_amount
            monday_duration += duration
            worksheet["W1"].value = monday_duration
        case 1:
            tuesday_amount += 1
            worksheet["V2"].value = tuesday_amount
            tuesday_duration += duration
            worksheet["W2"].value = tuesday_duration
        case 2:
            wednesday_amount += 1
            worksheet["V3"].value = wednesday_amount
            wednesday_duration += duration
            worksheet["W3"].value = wednesday_duration
        case 3:
            thursday_amount += 1
            worksheet["V4"].value = thursday_amount
            thursday_duration += duration
            worksheet["W4"].value = thursday_duration
        case 4:
            friday_amount += 1
            worksheet["V5"].value = friday_amount
            friday_duration += duration
            worksheet["W5"].value = friday_duration
        case 5:
            saturday_amount += 1
            worksheet["V6"].value = saturday_amount
            saturday_duration += duration
            worksheet["W6"].value = saturday_duration
        case 6:
            sunday_amount += 1
            worksheet["V7"].value = sunday_amount
            sunday_duration += duration
            worksheet["W7"].value = sunday_duration

    timer_time, break_time = 0, 0
    workbook.save(data_file)
    print("Weekday saved.")


def autopct_format(values):
    def my_format(pct):
        total = sum(values)
        val = int(round(pct*total/100.0))
        return "{v:d}".format(v=val)
    return my_format


def create_weekday_graph(day_duration_list, day_name_list):
    global pie_color_1, pie_color_2, pie_color_3, pie_color_4, pie_color_5, pie_color_6, pie_color_7
    non_zero_durations = [duration for duration in day_duration_list if duration != 0]
    non_zero_names = [name for name, duration in zip(day_name_list, day_duration_list) if duration != 0]

    fig2, ax2 = plt.subplots()
    ax2.pie(non_zero_durations, labels=non_zero_names, autopct=autopct_format(non_zero_durations), 
           colors=[pie_color_1, pie_color_2, pie_color_3, pie_color_4, pie_color_5, pie_color_6, pie_color_7], 
           textprops={"fontsize": pie_font_size, "family": pie_font_family, "color": font_color}, counterclock=False, startangle=90)
    fig2.set_size_inches(graph_width/100, graph_height/100, forward=True)
    fig2.set_facecolor(graph_bg_color)
    ax2.tick_params(colors="white")
    ax2.set_facecolor(graph_fg_color)
    ax2.set_title("Time Spent by Weekday", color=font_color)
    ax2.spines["top"].set_color(spine_color)
    ax2.spines["bottom"].set_color(spine_color)
    ax2.spines["left"].set_color(spine_color)
    ax2.spines["right"].set_color(spine_color)

    weekday_frame = FigureCanvasTkAgg(fig2, master=statistics_frame)

    weekday_graph = weekday_frame.get_tk_widget()
    weekday_graph.grid(row=0, column=1, padx=10, pady=10)
    weekday_graph.config(highlightbackground=frame_border_color, highlightthickness=2, background=frame_color)


def collect_data():
    global data_amount, date_list, duration_list, goal_amount, total_duration, day_duration_list, day_name_list
    global monday_amount, tuesday_amount, wednesday_amount, thursday_amount, friday_amount, saturday_amount, sunday_amount
    global monday_duration, tuesday_duration, wednesday_duration, thursday_duration, friday_duration, saturday_duration, sunday_duration
    global color, default_color, pie_color_1, pie_color_2, pie_color_3, pie_color_4, pie_color_5, pie_color_6, pie_color_7

    data_amount = int(worksheet["Z1"].value)

    monday_amount = int(worksheet["V1"].value)
    tuesday_amount = int(worksheet["V2"].value)
    wednesday_amount = int(worksheet["V3"].value)
    thursday_amount = int(worksheet["V4"].value)
    friday_amount = int(worksheet["V5"].value)
    saturday_amount = int(worksheet["V6"].value)
    sunday_amount = int(worksheet["V7"].value)

    monday_duration = int(worksheet["W1"].value)
    tuesday_duration = int(worksheet["W2"].value)
    wednesday_duration = int(worksheet["W3"].value)
    thursday_duration = int(worksheet["W4"].value)
    friday_duration = int(worksheet["W5"].value)
    saturday_duration = int(worksheet["W6"].value)
    sunday_duration = int(worksheet["W7"].value)

    goal_amount = int(worksheet["R1"].value)

    color = worksheet["T1"].value
    if color == None:
        color = "Orange"
    default_color = ctk.StringVar(value=color)
    pie_colors = {"Orange": [pie_color_orange_1, pie_color_orange_2, pie_color_orange_3, pie_color_orange_4, pie_color_orange_5, pie_color_orange_6, pie_color_orange_7], 
                  "Green": [pie_color_green_1, pie_color_green_2, pie_color_green_3, pie_color_green_4, pie_color_green_5, pie_color_green_6, pie_color_green_7], 
                  "Blue": [pie_color_blue_1, pie_color_blue_2, pie_color_blue_3, pie_color_blue_4, pie_color_blue_5, pie_color_blue_6, pie_color_blue_7]}
    pie_color_1, pie_color_2, pie_color_3, pie_color_4, pie_color_5, pie_color_6, pie_color_7 = pie_colors[color]


    day_duration_list = [monday_duration, tuesday_duration, wednesday_duration, thursday_duration, friday_duration, saturday_duration, sunday_duration]
    day_name_list = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    create_weekday_graph(day_duration_list, day_name_list)

    for data in range(2, data_amount + 2):
        if "/" in str(worksheet["B" + str(data)].value):
            date_list.append(datetime.datetime.strptime(str(worksheet["B" + str(data)].value).split(" ")[0], "%d/%m/%Y").date())
        elif "-" in str(worksheet["B" + str(data)].value):
            date_list.append(datetime.datetime.strptime(str(worksheet["B" + str(data)].value).split(" ")[0], "%Y-%m-%d").date())
        duration_list.append(round(worksheet["C" + str(data)].value))

    total_duration = sum(duration_list)

    create_time_spent_graph(date_list, duration_list)
    print("Data collected.")
    
if os.path.isfile(data_file):
    workbook = op.load_workbook(data_file)
    worksheet = workbook.active

    collect_data()
    print("File loaded")
else:
    workbook = op.Workbook()
    worksheet = workbook.active

    data_amount, goal_amount = 0, 0

    monday_amount, tuesday_amount, wednesday_amount, thursday_amount, friday_amount, saturday_amount, sunday_amount = 0, 0, 0, 0, 0, 0, 0

    monday_duration, tuesday_duration, wednesday_duration, thursday_duration, friday_duration, saturday_duration, sunday_duration = 0, 0, 0, 0, 0, 0, 0

    workbook.save(data_file)
    print("New file created")
    customize_excel(worksheet)


#-----------------------------------------------------------------------------TIMER-----------------------------------------------------------------------------#
def timer_mechanism():
    global timer_running, break_running, start_time
    global timer_btn, break_btn
    if not timer_running:
        timer_running = True
        break_running = False
        timer_btn.configure(text="Stop")
        break_btn.configure(text="Start")
        update_time()
    elif timer_running:
        timer_running = False
        timer_btn.configure(text="Start")
    if start_time == "":
        start_time = datetime.datetime.now()

def update_time():
    global timer_running, timer_time, time_display_label

    if timer_running:
        timer_time += 1
        time_display_label.configure(text=str(datetime.timedelta(seconds=timer_time)))
        update_slider(timer_time)
        WINDOW.after(1000, update_time)

def break_mechanism():
    global break_running, timer_running
    global break_btn, timer_btn
    if not break_running:
        break_running = True
        timer_running = False
        break_btn.configure(text="Stop")
        timer_btn.configure(text="Start")
        update_break_time()
    elif break_running:
        break_running = False
        break_btn.configure(text="Start")

def update_break_time():
    global break_running, break_time, break_display_label

    if break_running:
        break_time += 1
        break_display_label.configure(text=str(datetime.timedelta(seconds=break_time)))
        WINDOW.after(1000, update_break_time)


#------------------------------------------------------------------------------DATA-----------------------------------------------------------------------------#
def save_data():
    global data_amount, duration_list, date_list, goal_amount, time_studied_label, total_duration, times_studied_label, progressbar
    global timer_running, timer_time, start_time, timer_btn, timer_label
    global break_running, break_time, break_btn, break_label

    progressbar.set(0)

    if timer_time < 60:
        print("No data to save.")
        return
    timer_running, break_running = False, False

    duration = calculate_duration()

    data_amount += 1
    worksheet["Z1"].value = int(data_amount)

    stop_time = datetime.datetime.now()

    worksheet["A" + str((data_amount + 1))].value = start_time.strftime("%d/%m/%Y %H:%M")
    worksheet["B" + str((data_amount + 1))].value = stop_time.strftime("%d/%m/%Y %H:%M")
    worksheet["C" + str((data_amount + 1))].value = duration
    worksheet["D" + str((data_amount + 1))].value = break_time/60

    timer_btn.configure(text="Start")
    break_btn.configure(text="Start")
    time_display_label.configure(text="0:00:00")
    break_display_label.configure(text="0:00:00")
    start_time = ""
    workbook.save(data_file)

    if timer_time/60 >= goal:
        progressbar.set(1)
        goal_amount += 1
        worksheet["R1"].value = goal_amount
        times_studied_label.configure(text=goal_amount)

    print("Data saved.")
    save_weekday()
    collect_data()
    time_studied_label.configure(text=total_duration)


def reset_data():
    global data_amount, duration_list, date_list, goal_amount, total_duration, times_studied_label, time_studied_label
    global timer_time, timer_running, time_display_label
    global break_time, break_running, break_display_label
    global monday_duration, tuesday_duration, wednesday_duration, thursday_duration, friday_duration, saturday_duration, sunday_duration

    data_amount = 0
    del workbook[workbook.active.title]
    workbook.create_sheet()
    worksheet = workbook.active

    timer_running, break_running = False, False
    timer_time, break_time = 0, 0
    time_display_label.configure(text="0:00:00")
    break_display_label.configure(text="0:00:00")

    worksheet["Z1"].value = int(data_amount)
    workbook.save(data_file)
    monday_duration, tuesday_duration, wednesday_duration, thursday_duration, friday_duration, saturday_duration, sunday_duration = 0, 0, 0, 0, 0, 0, 0
    day_duration_list = [monday_duration, tuesday_duration, wednesday_duration, thursday_duration, friday_duration, saturday_duration, sunday_duration]
    day_name_list = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    goal_amount, total_duration = 0, 0
    time_studied_label.configure(text="0")
    times_studied_label.configure(text="0")
    duration_list.clear()
    date_list.clear()

    print("Data reset.")
    create_time_spent_graph(date_list, duration_list)
    create_weekday_graph(day_duration_list, day_name_list)
    customize_excel(worksheet)


def save_on_quit():
    global color
    global timer_time

    worksheet["T1"].value = color
    print("Color saved.")
    
    if timer_time > 0:
        save_data()
        print("Data saved on exit.")
    else: print("Quit.")

    workbook.save(data_file)
    WINDOW.destroy()

def to_main():
    main_frame.grid(column=2, row=0, padx=main_frame_pad_x)
    main_frame.grid_propagate(False)
    statistics_frame.grid_forget()
    settings_frame.grid_forget()

def to_statistics():
    statistics_frame.grid(column=2, row=0, padx=main_frame_pad_x)
    statistics_frame.grid_propagate(False)
    main_frame.grid_forget()
    settings_frame.grid_forget()

def to_settings():
    settings_frame.grid(column=2, row=0, padx=main_frame_pad_x)
    settings_frame.grid_propagate(False)
    main_frame.grid_forget()
    statistics_frame.grid_forget()

def get_goal():
    global goal, goal_dropdown
    x = 0
    choice = goal_dropdown.get()
    if "hour" in choice:
        x += int(choice.split(" ")[0]) * 60
    if "minutes" in choice and "hour" in choice:
        x += int(choice.split(", ")[1].removesuffix(" minutes"))
    if "hour" not in choice:
        x += int(choice.split(" ")[0])
    goal = x


def update_slider(timer_time):
    global progressbar, goal
    if goal == 0:
        goal = 60
    if (timer_time/60) < goal:
        progressbar.set((timer_time/60)/goal)


def load_color(color, widget_list, progressbar):
    global graph_color, pie_color_1, pie_color_2, pie_color_3, pie_color_4, pie_color_5, pie_color_6, pie_color_7
    pie_colors = {"Orange": [pie_color_orange_1, pie_color_orange_2, pie_color_orange_3, pie_color_orange_4, pie_color_orange_5, pie_color_orange_6, pie_color_orange_7], 
                  "Green": [pie_color_green_1, pie_color_green_2, pie_color_green_3, pie_color_green_4, pie_color_green_5, pie_color_green_6, pie_color_green_7], 
                  "Blue": [pie_color_blue_1, pie_color_blue_2, pie_color_blue_3, pie_color_blue_4, pie_color_blue_5, pie_color_blue_6, pie_color_blue_7]}
    pie_color_1, pie_color_2, pie_color_3, pie_color_4, pie_color_5, pie_color_6, pie_color_7 = pie_colors[color]
    highlight_colors = {"Orange": orange_highlight_color, "Green": green_highlight_color, "Blue": blue_highlight_color}
    highlight_color = highlight_colors[color]
    colors = {"Orange": orange_button_color, "Green": green_button_color, "Blue": blue_button_color}
    color = colors[color]
    graph_color = color
    change_color(color, highlight_color, widget_list, progressbar)

    
def change_color(color, highlight_color, widget_list, progressbar):
    global date_list, duration_list, day_duration_list, day_name_list
    colors = {orange_button_color: "Orange", green_button_color: "Green", blue_button_color: "Blue"}
    worksheet["T1"].value = colors[color]
    workbook.save(data_file)
    collect_data()
    for widget in widget_list:
        widget.configure(fg_color=color, hover_color=highlight_color)
    progressbar.configure(progress_color = color)
    
def set_color(widget):
    global color, graph_color, pie_color_1, pie_color_2, pie_color_3, pie_color_4, pie_color_5, pie_color_6, pie_color_7, day_duration_list, day_name_list
    worksheet["T1"].value = color
    workbook.save(data_file)
    print("Color saved.")
    color = widget.get()
    colors = {"Orange": orange_button_color, "Green": green_button_color, "Blue": blue_button_color}
    highlight_colors = {"Orange": orange_highlight_color, "Green": green_highlight_color, "Blue": blue_highlight_color}
    pie_colors = {"Orange": [pie_color_orange_1, pie_color_orange_2, pie_color_orange_3, pie_color_orange_4, pie_color_orange_5, pie_color_orange_6, pie_color_orange_7], 
                  "Green": [pie_color_green_1, pie_color_green_2, pie_color_green_3, pie_color_green_4, pie_color_green_5, pie_color_green_6, pie_color_green_7], 
                  "Blue": [pie_color_blue_1, pie_color_blue_2, pie_color_blue_3, pie_color_blue_4, pie_color_blue_5, pie_color_blue_6, pie_color_blue_7]}
    pie_color_1, pie_color_2, pie_color_3, pie_color_4, pie_color_5, pie_color_6, pie_color_7 = pie_colors[color]
    c = colors[color]
    highlight_color = highlight_colors[color]
    graph_color = c
    change_color(c, highlight_color, widget_list, progressbar)
    
#------------------------------------------------------------------------------GUI------------------------------------------------------------------------------#
def change_focus(event):
    event.widget.focus_set()

tab_frame = ctk.CTkFrame(WINDOW, width=tab_frame_width, height=HEIGHT+((widget_padding_x+frame_padding)*2), fg_color=tab_frame_color)
tab_frame.grid(column=0, row=0)
tab_frame.pack_propagate(False)

goal_progress_frame = ctk.CTkFrame(main_frame, height=(HEIGHT-button_height*1.5), width=frame_width)
goal_progress_frame.grid(row=0, column=0)
goal_progress_frame.pack_propagate(False)

goal_frame = ctk.CTkFrame(goal_progress_frame, fg_color=frame_color, height=175, width=frame_width, corner_radius=10)
goal_frame.pack(padx=frame_padding, pady=frame_padding)
goal_frame.pack_propagate(False)

goal_label = ctk.CTkLabel(goal_frame, text="Goal", font=(font_family, font_size), text_color=font_color)
goal_label.place(anchor="nw", relx=0.05, rely=0.05)

goal_dropdown = ctk.CTkComboBox(goal_frame, values=["30 minutes", "1 hour", "1 hour, 30 minutes", "2 hours", "2 hours, 30 minutes", "3 hours", "3 hours, 30 minutes",
                                                     "4 hours", "4 hours, 30 minutes", "5 hours", "5 hours, 30 minutes", "6 hours"], variable=default_choice, 
                                                     state="readonly", width=200, height=30, dropdown_font=(font_family, int(font_size*0.75)),
                                                       font=(font_family, int(font_size)), fg_color=border_frame_color, button_color=border_frame_color)
goal_dropdown.place(anchor="center", relx=0.5, rely=0.45)
goal_btn = ctk.CTkButton(goal_frame, text="Save", font=(font_family, font_size), text_color=button_font_color, fg_color=button_color, hover_color=button_highlight_color,
                         height=button_height, command=get_goal)
goal_btn.place(anchor="s", relx=0.5, rely=0.9)

progress_frame = ctk.CTkFrame(goal_progress_frame, fg_color=frame_color, width=frame_width, corner_radius=10, height=100)
progress_frame.pack(padx=frame_padding, pady=frame_padding)
progress_frame.pack_propagate(False)
progress_label = ctk.CTkLabel(progress_frame, text="Progress", font=(font_family, int(font_size)), text_color=font_color)
progress_label.place(anchor="nw", relx=0.05, rely=0.05)
progressbar = ctk.CTkProgressBar(progress_frame, height=20, width=220, progress_color=button_color, fg_color=border_frame_color, corner_radius=10)
progressbar.place(anchor="center", relx=0.5, rely=0.65)
progressbar.set(0)

streak_frame = ctk.CTkFrame(goal_progress_frame, fg_color=frame_color, width=frame_width, corner_radius=10, height=220)
streak_frame.pack(padx=frame_padding, pady=frame_padding)
streak_label = ctk.CTkLabel(streak_frame, text="Streak", font=(font_family, int(font_size)), text_color=font_color)
streak_label.place(anchor="nw", relx=0.05, rely=0.05)
times_studied_text = ctk.CTkLabel(streak_frame, text="Goal\nreached", font=(font_family, int(font_size/1.25)), text_color=font_color)
times_studied_text.place(anchor="center", relx=0.3, rely=0.4)
times_studied_label = ctk.CTkLabel(streak_frame, text=goal_amount, font=(font_family, int(font_size*2.7)), text_color=font_color)
times_studied_label.place(anchor="center", relx=0.3, rely=0.6)
times_reached_label = ctk.CTkLabel(streak_frame, text="times", font=(font_family, int(font_size/1.25)), text_color=font_color)
times_reached_label.place(anchor="center", relx=0.3, rely=0.8)

time_studied_text = ctk.CTkLabel(streak_frame, text="Time\nstudied", font=(font_family, int(font_size/1.25)), text_color=font_color)
time_studied_text.place(anchor="center", relx=0.7, rely=0.4)
time_studied_label = ctk.CTkLabel(streak_frame, text=total_duration, font=(font_family, int(font_size*2.7)), text_color=font_color)
time_studied_label.place(anchor="center", relx=0.7, rely=0.6)
time_reached_label = ctk.CTkLabel(streak_frame, text="minutes", font=(font_family, int(font_size/1.25)), text_color=font_color)
time_reached_label.place(anchor="center", relx=0.7, rely=0.8)

timer_break_frame = ctk.CTkFrame(main_frame, height=(HEIGHT-button_height*1.5), width=frame_width)
timer_break_frame.grid(row=0, column=1)
timer_break_frame.propagate(False)

timer_frame = ctk.CTkFrame(timer_break_frame, fg_color=frame_color, corner_radius=10, width=frame_width, height=220)
timer_frame.pack(padx=frame_padding, pady=frame_padding)
timer_frame.pack_propagate(False)

timer_label = ctk.CTkLabel(timer_frame, text="Timer", font=(font_family, font_size), text_color=font_color)
timer_label.place(anchor="nw", relx=0.05, rely=0.05)
time_display_label = ctk.CTkLabel(timer_frame, text="0:00:00", font=(font_family, int(font_size*3)), text_color=font_color)
time_display_label.place(anchor="center", relx=0.5, rely=0.45)
timer_btn = ctk.CTkButton(timer_frame, text="Start", font=(font_family, font_size), fg_color=button_color, text_color=button_font_color,
                                 border_color=frame_border_color, hover_color=button_highlight_color, height=button_height, command=timer_mechanism)
timer_btn.place(anchor="s", relx=0.5, rely=0.9)

#BREAK TIMER UI ROW
break_frame = ctk.CTkFrame(timer_break_frame, fg_color=frame_color, corner_radius=10, width=frame_width, height=220)
break_frame.pack(padx=frame_padding, pady=frame_padding)
break_frame.pack_propagate(False)

break_label = ctk.CTkLabel(break_frame, text="Break", font=(font_family, font_size), text_color=font_color)
break_label.place(anchor="nw", relx=0.05, rely=0.05)
break_display_label = ctk.CTkLabel(break_frame, text="0:00:00", font=(font_family, int(font_size*3)), text_color=font_color)
break_display_label.place(anchor="center", relx=0.5, rely=0.45)
break_btn = ctk.CTkButton(break_frame, text="Start", font=(font_family, font_size), fg_color=button_color, text_color=button_font_color,
                                 border_color=frame_border_color, hover_color=button_highlight_color, height=button_height, command=break_mechanism)
break_btn.place(anchor="s", relx=0.5, rely=0.9)

#DATA UI ROW
data_frame = ctk.CTkFrame(main_frame, fg_color=frame_color, corner_radius=10, width=WIDTH-10, height=button_height*2)
data_frame.place(anchor="s", relx=0.5, rely=0.985)
data_frame.grid_propagate(False)
save_data_btn = ctk.CTkButton(data_frame, text="Save Data", font=(font_family, font_size), fg_color=button_color, text_color=button_font_color,
                               border_color=frame_border_color, hover_color=button_highlight_color, height=button_height, command=save_data, width=450)
save_data_btn.place(relx=0.5, anchor="center", rely=0.5)

#TABS
timer_tab = ctk.CTkFrame(tab_frame, width=tab_frame_width, height=tab_height*0.8, fg_color=tab_color)
timer_tab.pack(pady=tab_padding_y)
timer_tab_btn = ctk.CTkButton(timer_tab, text="Timer", font=(tab_font_family, 22*tab_height/55, tab_font_weight), text_color=font_color,
                                 fg_color=tab_color, width=int(tab_frame_width*0.95), height=int(tab_height*0.7), hover_color=tab_highlight_color, anchor="w", command=to_main)
timer_tab_btn.place(relx=0.5, rely=0.5, anchor="center")

statistics_tab = ctk.CTkFrame(tab_frame, width=tab_frame_width, height=tab_height*0.8, fg_color=tab_color)
statistics_tab.pack(pady=tab_padding_y)
statistics_btn = ctk.CTkButton(statistics_tab, text="Statistics", font=(tab_font_family, 22*tab_height/55, tab_font_weight), text_color=font_color,
                                 fg_color=tab_color, width=int(tab_frame_width*0.95), height=int(tab_height*0.7), hover_color=tab_highlight_color, anchor="w", command=to_statistics)
statistics_btn.place(relx=0.5, rely=0.5, anchor="center")

settings_tab = ctk.CTkFrame(tab_frame, width=tab_frame_width, height=tab_height*0.8, fg_color=tab_color)
settings_tab.place(relx=0.5, rely=1, anchor="s")
settings_btn = ctk.CTkButton(settings_tab, text="Settings", font=(tab_font_family, 22*tab_height/55, tab_font_weight), text_color=font_color,
                                 fg_color=tab_color, width=int(tab_frame_width*0.95), height=int(tab_height*0.7), hover_color=tab_highlight_color, anchor="w", command=to_settings)
settings_btn.place(relx=0.5, rely=0.5, anchor="center")

color_select_frame = ctk.CTkFrame(settings_frame, fg_color=frame_color, height=200, width=int(frame_width/1.25), corner_radius=10)
color_select_frame.grid(column=0, row=0, padx=frame_padding, pady=frame_padding)
color_select_frame.pack_propagate(False)
color_label = ctk.CTkLabel(color_select_frame, text="Color", font=(font_family, font_size), text_color=font_color)
color_label.place(anchor="nw", relx=0.05, rely=0.05)
color_dropdown = ctk.CTkComboBox(color_select_frame, values=["Orange", "Green", "Blue"], variable=default_color, state="readonly", width=150, height=30, 
                                 dropdown_font=(font_family, int(font_size*0.75)), font=(font_family, int(font_size)), fg_color=border_frame_color, button_color=border_frame_color)
color_dropdown.place(anchor="center", relx=0.5, rely=0.45)
color_btn = ctk.CTkButton(color_select_frame, text="Save", font=(font_family, font_size), text_color=button_font_color, fg_color=button_color, hover_color=button_highlight_color,
                         height=button_height, command=lambda: set_color(color_dropdown))
color_btn.place(anchor="s", relx=0.5, rely=0.9)

reset_btn_frame = ctk.CTkFrame(settings_frame, fg_color=tab_color)
reset_btn_frame.place(anchor="s", relx=0.5, rely=0.985)
reset_data_btn = ctk.CTkButton(reset_btn_frame, text="Reset Data", font=(font_family, font_size), fg_color=button_color, text_color=button_font_color,
                                border_color=frame_border_color, hover_color=button_highlight_color, height=button_height, command=reset_data, width=450)
reset_data_btn.pack()

widget_list = [goal_btn, timer_btn, break_btn, save_data_btn, color_btn, reset_data_btn]

load_color(color, widget_list, progressbar)

WINDOW.bind_all("<Button>", change_focus)

WINDOW.protocol("WM_DELETE_WINDOW", save_on_quit)

WINDOW.mainloop()