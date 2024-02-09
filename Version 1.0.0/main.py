import os
import random

import openpyxl as op
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import customtkinter as ctk
from matplotlib.ticker import MaxNLocator, FuncFormatter
from winotify import Notification

from Package import *



class App:
    def __init__(self):
        self.APPNAME = "Timer App"
        self.FILENAME = "Timer Data.xlsx"

        self._window_setup()
        self.initialize_variables()
        self.create_gui()
        self._file_setup()
        self.create_time_spent_graph()
        self.create_weekday_graph()
        self.WINDOW.protocol("WM_DELETE_WINDOW", self.save_on_quit)


    def create_gui(self):
        self._main_frame_gui_setup()
        self._tab_frames_gui_setup()
        self._secondary_frames_gui_setup()

        self._timer_gui_setup()
        self._break_gui_setup()
        self._save_data_gui()
        self._goal_gui_setup()
        self._progress_gui_setup()
        self._streak_gui_setup()
        self._settings_gui_setup()


    def _file_setup(self):
        self.local_folder = os.path.expandvars(rf"%APPDATA%\{self.APPNAME}")
        self.data_file = os.path.expandvars(rf"%APPDATA%\{self.APPNAME}\{self.FILENAME}")

        os.makedirs(self.local_folder, exist_ok=True)

        self.timer_manager = TimerManager(self, self.WINDOW)

        data_file_exists = os.path.isfile(self.data_file)

        if data_file_exists:
            self.workbook = op.load_workbook(self.data_file)
            self.worksheet = self.workbook.active

            self.data_manager = DataManager(self, self.timer_manager, self.workbook, self.worksheet)

            self.collect_data()
            self.update_streak_values()
            self.create_widget_list()
            self.data_manager.load_color()
            print("File loaded.")

        else:
            self.workbook = op.Workbook()
            self.worksheet = self.workbook.active

            self.workbook.save(self.data_file)

            self.data_manager = DataManager(self, self.timer_manager, self.workbook, self.worksheet)

            self.data_manager.initialize_new_file_variables()
            self.data_manager.customize_excel()

            self.create_widget_list()

            print("New file created.")


    def initialize_variables(self):
        self.default_choice = ctk.StringVar(value="1 hour")
        self.notification_limit_on = False
        self.goal = 60


    def _window_setup(self):
        self.WINDOW = ctk.CTk()
        self.WINDOW.geometry(str(WIDTH + BORDER_WIDTH + main_frame_pad_x + tab_frame_width) + "x" + str(HEIGHT+((widget_padding_x+frame_padding)*2)))
        self.WINDOW.title(self.APPNAME)
        self.WINDOW.configure(background=window_color)
        self.WINDOW.resizable(False, False)
        self.WINDOW.grid_propagate(False)


    def _main_frame_gui_setup(self):
        self.tab_frame = ctk.CTkFrame(self.WINDOW, width=tab_frame_width, height=HEIGHT+((widget_padding_x+frame_padding)*2), fg_color=tab_frame_color)
        self.tab_frame.grid(column=0, row=0)
        self.tab_frame.pack_propagate(False)

        self.main_frame = ctk.CTkFrame(self.WINDOW, fg_color=main_frame_color, height=HEIGHT+((widget_padding_x+frame_padding)*2), width=WIDTH, corner_radius=0)
        self.main_frame.grid(column=2, row=0, padx=main_frame_pad_x)
        self.main_frame.grid_propagate(False)

        self.statistics_frame = ctk.CTkFrame(self.WINDOW, fg_color=main_frame_color, height=HEIGHT+((widget_padding_x+frame_padding)*2), width=WIDTH, corner_radius=0)
        self.statistics_frame.grid(column=2, row=0, padx=main_frame_pad_x)

        self.settings_frame = ctk.CTkFrame(self.WINDOW, fg_color=main_frame_color, height=HEIGHT+((widget_padding_x+frame_padding)*2), width=WIDTH, corner_radius=0)
        self.settings_frame.grid(column=2, row=0, padx=main_frame_pad_x)
        self.settings_frame.grid_forget()

        self.achievements_frame = ctk.CTkFrame(self.WINDOW, fg_color=main_frame_color, height=HEIGHT+((widget_padding_x+frame_padding)*2), width=WIDTH, corner_radius=0)
        self.achievements_frame.grid(column=2, row=0, padx=main_frame_pad_x)

        self.history_frame = ctk.CTkFrame(self.WINDOW, fg_color=main_frame_color, height=HEIGHT+((widget_padding_x+frame_padding)*2), width=WIDTH, corner_radius=0)
        self.history_frame.grid(column=2, row=0, padx=main_frame_pad_x)

        self.forget_and_propagate([self.statistics_frame, self.settings_frame, self.achievements_frame, self.history_frame])


    def _tab_frames_gui_setup(self):
        self.timer_tab = ctk.CTkFrame(self.tab_frame, width=tab_frame_width, height=tab_height*0.8, fg_color=tab_color)
        self.timer_tab.pack(pady=tab_padding_y)
        self.timer_tab_button = ctk.CTkButton(self.timer_tab, text="Timer", font=(tab_font_family, 22*tab_height/50, tab_font_weight), text_color=font_color,
                                              fg_color=tab_color, width=int(tab_frame_width*0.95), height=int(tab_height*0.7), hover_color=tab_highlight_color, 
                                              anchor="w", command=lambda: self.switch_tab("main"))
        self.timer_tab_button.place(relx=0.5, rely=0.5, anchor="center")

        self.statistics_tab = ctk.CTkFrame(self.tab_frame, width=tab_frame_width, height=tab_height*0.8, fg_color=tab_color)
        self.statistics_tab.pack(pady=tab_padding_y)
        self.statistics_tab_button = ctk.CTkButton(self.statistics_tab, text="Statistics", font=(tab_font_family, 22*tab_height/50, tab_font_weight), text_color=font_color,
                                                   fg_color=tab_color, width=int(tab_frame_width*0.95), height=int(tab_height*0.7), hover_color=tab_highlight_color, 
                                                   anchor="w", command=lambda: self.switch_tab("statistics"))
        self.statistics_tab_button.place(relx=0.5, rely=0.5, anchor="center")

        self.achievements_tab = ctk.CTkFrame(self.tab_frame, width=tab_frame_width, height=tab_height*0.8, fg_color=tab_color)
        self.achievements_tab.pack(pady=tab_padding_y)
        self.achievements_tab_button = ctk.CTkButton(self.achievements_tab, text="Achievements", font=(tab_font_family, 22*tab_height/50, tab_font_weight), text_color=font_color,
                                                     fg_color=tab_color, width=int(tab_frame_width*0.95), height=int(tab_height*0.7), hover_color=tab_highlight_color, 
                                                     anchor="w", command=lambda: self.switch_tab("achievements"))
        self.achievements_tab_button.place(relx=0.5, rely=0.5, anchor="center")

        self.history_tab = ctk.CTkFrame(self.tab_frame, width=tab_frame_width, height=tab_height*0.8, fg_color=tab_color)
        self.history_tab.pack(pady=tab_padding_y)
        self.history_tab_button = ctk.CTkButton(self.history_tab, text="History", font=(tab_font_family, 22*tab_height/50, tab_font_weight), text_color=font_color,
                                                fg_color=tab_color, width=int(tab_frame_width*0.95), height=int(tab_height*0.7), hover_color=tab_highlight_color, 
                                                anchor="w", command=lambda: self.switch_tab("history"))
        self.history_tab_button.place(relx=0.5, rely=0.5, anchor="center")

        self.settings_tab = ctk.CTkFrame(self.tab_frame, width=tab_frame_width, height=tab_height*0.8, fg_color=tab_color)
        self.settings_tab.place(relx=0.5, rely=1, anchor="s")
        settings_tab_button = ctk.CTkButton(self.settings_tab, text="Settings", font=(tab_font_family, 22*tab_height/50, tab_font_weight), text_color=font_color,
                                 fg_color=tab_color, width=int(tab_frame_width*0.95), height=int(tab_height*0.7), hover_color=tab_highlight_color, anchor="w", command=lambda: self.switch_tab("settings"))
        settings_tab_button.place(relx=0.5, rely=0.5, anchor="center")


    def _secondary_frames_gui_setup(self):
        self.timer_break_frame = ctk.CTkFrame(self.main_frame, height=(HEIGHT-button_height*1.5), width=frame_width)
        self.timer_break_frame.grid(row=0, column=1)
        self.timer_break_frame.pack_propagate(False)

        self.goal_progress_frame = ctk.CTkFrame(self.main_frame, height=(HEIGHT-button_height*1.5), width=frame_width)
        self.goal_progress_frame.grid(row=0, column=0)
        self.goal_progress_frame.pack_propagate(False)

        self.subject_pomodoro_frame = ctk.CTkFrame(self.main_frame, height=(HEIGHT-button_height*1.5), width=frame_width)
        self.subject_pomodoro_frame.grid(row=0, column=2)
        self.subject_pomodoro_frame.pack_propagate(False)


    def _goal_gui_setup(self):
        goal_frame = ctk.CTkFrame(self.goal_progress_frame, fg_color=frame_color, height=175, width=frame_width, corner_radius=10)
        goal_frame.pack(padx=frame_padding, pady=frame_padding)
        goal_frame.pack_propagate(False)

        goal_label = ctk.CTkLabel(goal_frame, text="Goal", font=(font_family, font_size), text_color=font_color)
        goal_label.place(anchor="nw", relx=0.05, rely=0.05)

        self.goal_dropdown = ctk.CTkComboBox(goal_frame, values=["1 minutes", "30 minutes", "1 hour", "1 hour, 30 minutes", "2 hours", "2 hours, 30 minutes", "3 hours", "3 hours, 30 minutes",
                                                            "4 hours", "4 hours, 30 minutes", "5 hours", "5 hours, 30 minutes", "6 hours"], variable=self.default_choice, 
                                                            state="readonly", width=200, height=30, dropdown_font=(font_family, int(font_size*0.75)),
                                                            font=(font_family, int(font_size)), fg_color=border_frame_color, button_color=border_frame_color)
        self.goal_dropdown.place(anchor="center", relx=0.5, rely=0.45)

        self.goal_button = ctk.CTkButton(goal_frame, text="Save", font=(font_family, font_size), text_color=button_font_color, fg_color=button_color, hover_color=button_highlight_color,
                                height=button_height, command=self.set_goal)
        self.goal_button.place(anchor="s", relx=0.5, rely=0.9)


    def _progress_gui_setup(self):
        progress_frame = ctk.CTkFrame(self.goal_progress_frame, fg_color=frame_color, width=frame_width, corner_radius=10, height=100)
        progress_frame.pack(padx=frame_padding, pady=frame_padding)
        progress_frame.pack_propagate(False)

        progress_label = ctk.CTkLabel(progress_frame, text="Progress", font=(font_family, int(font_size)), text_color=font_color)
        progress_label.place(anchor="nw", relx=0.05, rely=0.05)

        self.progressbar = ctk.CTkProgressBar(progress_frame, height=20, width=220, progress_color=button_color, fg_color=border_frame_color, corner_radius=10)
        self.progressbar.place(anchor="center", relx=0.5, rely=0.65)
        self.progressbar.set(0)


    def _streak_gui_setup(self):
        streak_frame = ctk.CTkFrame(self.goal_progress_frame, fg_color=frame_color, width=frame_width, corner_radius=10, height=220)
        streak_frame.pack(padx=frame_padding, pady=frame_padding)

        streak_label = ctk.CTkLabel(streak_frame, text="Streak", font=(font_family, int(font_size)), text_color=font_color)
        streak_label.place(anchor="nw", relx=0.05, rely=0.05)
        
        times_studied_text = ctk.CTkLabel(streak_frame, text="Goal\nreached", font=(font_family, int(font_size/1.25)), text_color=font_color)
        times_studied_text.place(anchor="center", relx=0.3, rely=0.4)
        self.times_goal_reached = ctk.CTkLabel(streak_frame, text=0, font=(font_family, int(font_size*2.7)), text_color=font_color)
        self.times_goal_reached.place(anchor="center", relx=0.3, rely=0.6)
        times_reached_label = ctk.CTkLabel(streak_frame, text="times", font=(font_family, int(font_size/1.25)), text_color=font_color)
        times_reached_label.place(anchor="center", relx=0.3, rely=0.8)

        duration_studied_text = ctk.CTkLabel(streak_frame, text="Time\nstudied", font=(font_family, int(font_size/1.25)), text_color=font_color)
        duration_studied_text.place(anchor="center", relx=0.7, rely=0.4)
        self.streak_duration = ctk.CTkLabel(streak_frame, text=0, font=(font_family, int(font_size*2.7)), text_color=font_color)
        self.streak_duration.place(anchor="center", relx=0.7, rely=0.6)
        duration_minute_label = ctk.CTkLabel(streak_frame, text="minutes", font=(font_family, int(font_size/1.25)), text_color=font_color)
        duration_minute_label.place(anchor="center", relx=0.7, rely=0.8)


    def _timer_gui_setup(self):
        timer_frame = ctk.CTkFrame(self.timer_break_frame, fg_color=frame_color, corner_radius=10, width=frame_width, height=220)
        timer_frame.pack(padx=frame_padding, pady=frame_padding)
        timer_frame.pack_propagate(False)

        timer_label = ctk.CTkLabel(timer_frame, text="Timer", font=(font_family, font_size), text_color=font_color)
        timer_label.place(anchor="nw", relx=0.05, rely=0.05)
        self.time_display_label = ctk.CTkLabel(timer_frame, text="0:00:00", font=(font_family, int(font_size*3)), text_color=font_color)
        self.time_display_label.place(anchor="center", relx=0.5, rely=0.45)
        self.timer_button = ctk.CTkButton(timer_frame, text="Start", font=(font_family, font_size), fg_color=button_color, text_color=button_font_color,
                                        border_color=frame_border_color, hover_color=button_highlight_color, height=button_height, command=self.timer_mechanism)
        self.timer_button.place(anchor="s", relx=0.5, rely=0.9)
        
    
    def _break_gui_setup(self):
        break_frame = ctk.CTkFrame(self.timer_break_frame, fg_color=frame_color, corner_radius=10, width=frame_width, height=220)
        break_frame.pack(padx=frame_padding, pady=frame_padding)
        break_frame.pack_propagate(False)

        break_label = ctk.CTkLabel(break_frame, text="Break", font=(font_family, font_size), text_color=font_color)
        break_label.place(anchor="nw", relx=0.05, rely=0.05)
        self.break_display_label = ctk.CTkLabel(break_frame, text="0:00:00", font=(font_family, int(font_size*3)), text_color=font_color)
        self.break_display_label.place(anchor="center", relx=0.5, rely=0.45)
        self.break_button = ctk.CTkButton(break_frame, text="Start", font=(font_family, font_size), fg_color=button_color, text_color=button_font_color,
                                        border_color=frame_border_color, hover_color=button_highlight_color, height=button_height, command=self.break_mechanism)
        self.break_button.place(anchor="s", relx=0.5, rely=0.9)

    def _save_data_gui(self):
        self.data_frame = ctk.CTkFrame(self.main_frame, fg_color=frame_color, corner_radius=10, width=WIDTH-10, height=button_height*2)
        self.data_frame.place(anchor="s", relx=0.5, rely=0.985)
        self.data_frame.grid_propagate(False)
        self.save_data_button = ctk.CTkButton(self.data_frame, text="Save Data", font=(font_family, font_size), fg_color=button_color, text_color=button_font_color,
                                    border_color=frame_border_color, hover_color=button_highlight_color, height=button_height, width=450, command=self.save_data)
        self.save_data_button.place(relx=0.5, anchor="center", rely=0.5)


    def _settings_gui_setup(self):
        color_select_frame = ctk.CTkFrame(self.settings_frame, fg_color=frame_color, height=200, width=int(frame_width/1.25), corner_radius=10)
        color_select_frame.grid(column=0, row=0, padx=frame_padding, pady=frame_padding)
        color_label = ctk.CTkLabel(color_select_frame, text="Color", font=(font_family, font_size), text_color=font_color)
        color_label.place(anchor="nw", relx=0.05, rely=0.05)
        self.color_dropdown = ctk.CTkComboBox(color_select_frame, values=["Orange", "Green", "Blue"], state="readonly", width=150, height=30,
                                         dropdown_font=(font_family, int(font_size*0.75)), font=(font_family, int(font_size)), fg_color=border_frame_color, button_color=border_frame_color)
        self.color_dropdown.place(anchor="center", relx=0.5, rely=0.45)
        self.color_button = ctk.CTkButton(color_select_frame, text="Save", font=(font_family, font_size), text_color=button_font_color, fg_color=button_color, hover_color=button_highlight_color,
                                     height=button_height, command=lambda: self.data_manager.set_color(self.color_dropdown))
        self.color_button.place(anchor="s", relx=0.5, rely=0.9)

        reset_frame = ctk.CTkFrame(self.settings_frame, fg_color=tab_color)
        reset_frame.place(anchor="s", relx=0.5, rely=0.985)
        self.reset_data_button = ctk.CTkButton(reset_frame, text="Reset Data", font=(font_family, font_size), fg_color=button_color, text_color=button_font_color,
                                        border_color=frame_border_color, hover_color=button_highlight_color, height=button_height, command=self.reset_data, width=450)
        self.reset_data_button.pack()


    def _subject_gui_setup(self):
        subject_frame = ctk.CTkFrame(self.subject_pomodoro_frame, fg_color=frame_color, height=175, width=frame_width, corner_radius=10)
        subject_frame.pack(padx=frame_padding, pady=frame_padding)
        subject_label = ctk.CTkLabel(subject_frame, text="Subject", font=(font_family, font_size), text_color=font_color)
        subject_label.place(anchor="nw", relx=0.05, rely=0.05)
        subject_selection = ctk.CTkComboBox(subject_frame, values=["Mathematics", "Science", "Literature", "History", "Geography", "Language Arts", "Foreign Languages", "Social Studies",
                                                                "Economics", "Computer Science", "Psychology", "Philosophy", "Art", "Music", "Physical Education", "Other"], 
                                                            state="readonly", width=200, height=30, dropdown_font=(font_family, int(font_size*0.75)),
                                                            font=(font_family, int(font_size)), fg_color=border_frame_color, button_color=border_frame_color)
        subject_selection.place(anchor="center", relx=0.5, rely=0.45)
        subject_btn = ctk.CTkButton(subject_frame, text="Save", font=(font_family, font_size), text_color=button_font_color, fg_color=button_color, hover_color=button_highlight_color,
                                height=button_height)
        subject_btn.place(anchor="s", relx=0.5, rely=0.9)

    def create_time_spent_graph(self):
        data = {"Date": self.data_manager.date_list, "Duration": self.data_manager.duration_list}
        df = pd.DataFrame(data)
        grouped_data = df.groupby("Date")["Duration"].sum().reset_index()
        fig1, ax = plt.subplots()
        ax.bar(grouped_data["Date"], grouped_data["Duration"], color=self.data_manager.graph_color)
        ax.set_title("Duration of Study Sessions by Date", color=font_color)
        ax.tick_params(colors="white")
        ax.set_facecolor(graph_fg_color)
        fig1.set_facecolor(graph_bg_color)
        ax.spines["top"].set_color(spine_color)
        ax.spines["bottom"].set_color(spine_color)
        ax.spines["left"].set_color(spine_color)
        ax.spines["right"].set_color(spine_color)
        fig1.set_size_inches(graph_width/100, graph_height/100, forward=True)
        ax.tick_params(axis='x', labelrotation = 45)

        def _format_func(value, tick_number):
            return f"{int(value)} m"
        
        plt.gca().yaxis.set_major_formatter(FuncFormatter(_format_func))
        date_format = mdates.DateFormatter("%d/%m")
        ax.xaxis.set_major_formatter(date_format)
        ax.xaxis.set_major_locator(MaxNLocator(integer=True, prune='both'))
        time_spent_frame = FigureCanvasTkAgg(fig1, master=self.statistics_frame)
        plt.subplots_adjust(bottom=0.2)

        time_spent_graph = time_spent_frame.get_tk_widget()
        time_spent_graph.grid(row=0, column=0, padx=10, pady=10)
        time_spent_graph.config(highlightbackground=frame_border_color, highlightthickness=2, background=frame_color)

        self.data_manager.clear_graph_lists()


    def create_weekday_graph(self):
        day_duration_list = self.data_manager.collect_day_data()
        non_zero_durations = [duration for duration in day_duration_list if duration != 0]
        non_zero_names = [name for name, duration in zip(self.data_manager.day_name_list, day_duration_list) if duration != 0]

        def _autopct_format(values):
            def _my_format(pct):
                total = sum(values)
                val = int(round(pct*total/100.0))
                return "{v:d} m".format(v=val)
            return _my_format

        fig, ax = plt.subplots()
        ax.pie(non_zero_durations, labels=non_zero_names, autopct=_autopct_format(non_zero_durations), 
            colors=self.data_manager.pie_colors, 
            textprops={"fontsize": pie_font_size, "family": pie_font_family, "color": font_color}, counterclock=False, startangle=90)
        fig.set_size_inches(graph_width/100, graph_height/100, forward=True)
        fig.set_facecolor(graph_bg_color)
        ax.tick_params(colors="white")
        ax.set_facecolor(graph_fg_color)
        ax.set_title("Duration of Study Sessions by Day of the Week", color=font_color)
        ax.spines["top"].set_color(spine_color)
        ax.spines["bottom"].set_color(spine_color)
        ax.spines["left"].set_color(spine_color)
        ax.spines["right"].set_color(spine_color)

        weekday_frame = FigureCanvasTkAgg(fig, master=self.statistics_frame)

        weekday_graph = weekday_frame.get_tk_widget()
        weekday_graph.grid(row=0, column=1, padx=10, pady=10)
        weekday_graph.config(highlightbackground=frame_border_color, highlightthickness=2, background=frame_color)

    
    def forget_and_propagate(self, list: list):
        for item in list:
            item.grid_forget()
            item.grid_propagate(False)


    def switch_tab(self, tab = str):
        tabs = {
            "main": self.main_frame,
            "statistics": self.statistics_frame,
            "settings": self.settings_frame,
            "achievements": self.achievements_frame,
            "history": self.history_frame
        }

        tab_list = [self.main_frame, self.statistics_frame, self.settings_frame, self.achievements_frame, self.history_frame]
        tab_list.remove(tabs[tab])

        def forget_tabs():
            for frame in tab_list:
                frame.grid_forget()
        forget_tabs()

        tabs[tab].grid(column=2, row=0, padx=main_frame_pad_x)
        tabs[tab].grid_propagate(False)


    def timer_mechanism(self):
        self.timer_manager.timer_mechanism(self.timer_button, self.break_button, self.time_display_label)
        
        #Get start time only at the start of timer
        if self.timer_manager.timer_time <= 1:
            self.data_manager.get_start_time()


    def reset_gui_values(self):
        self.times_goal_reached.configure(text=0)
        self.streak_duration.configure(text=0)
        self.progressbar.set(0)
        self.reset_timers()


    def reset_timers(self):
        self.timer_button.configure(text="Start")
        self.break_button.configure(text="Start")
        self.time_display_label.configure(text="0:00:00")
        self.break_display_label.configure(text="0:00:00") 


    def set_goal(self):
        x = 0
        choice = self.goal_dropdown.get()

        #Make an int out of a string e.g. "1 hour, 30 minutes"
        if "hour" in choice:
            x += int(choice.split(" ")[0]) * 60
        if "minutes" in choice and "hour" in choice:
            x += int(choice.split(", ")[1].removesuffix(" minutes"))
        if "hour" not in choice:
            x += int(choice.split(" ")[0])
        self.goal = x


    def reach_goal(self, timer_time: int):
        time_in_minutes = timer_time / 60
        if time_in_minutes < self.goal:
            self.progressbar.set(time_in_minutes/self.goal)
        elif time_in_minutes >= self.goal and not self.notification_limit_on:
            self.progressbar.set(1)

            message = random.choice(["Congratulations! You've reached your study goal. Take a well-deserved break and recharge!", "Study session complete! Great job on reaching your goal. Time for a quick break!",
                                     "You did it! Study session accomplished. Treat yourself to a moment of relaxation!", "Well done! You've met your study goal. Now, take some time to unwind and reflect on your progress.",
                                     "Study session over! You've achieved your goal. Reward yourself with a brief pause before your next task.", "Goal achieved! Take a breather and pat yourself on the back for your hard work.",
                                     "Mission accomplished! You've hit your study target. Enjoy a short break before diving back in.", "Study session complete. Nicely done! Use this time to relax and rejuvenate before your next endeavor.",
                                     "You've reached your study goal! Treat yourself to a well-deserved break. You've earned it!", "Goal achieved! Take a moment to celebrate your success. Your dedication is paying off!"])
            self.send_notification("Study Goal Reached", message)


    def update_streak_values(self):
        self.times_goal_reached.configure(text=self.data_manager.goal_amount)
        self.streak_duration.configure(text=self.data_manager.total_duration)

    
    def break_mechanism(self):
        self.timer_manager.break_mechanism(self.break_button, self.timer_button, self.break_display_label)


    def collect_data(self):
        self.data_manager.collect_data()
        self.data_manager.data_to_variable()


    def save_data(self):
        time_in_minutes = self.timer_manager.timer_time / 60

        #Only be able to save if time is higher than 1m
        if time_in_minutes >= 1:
            if time_in_minutes >= self.goal:
                self.data_manager.increase_goal_streak()

            self.data_manager.save_data()

            self.collect_data()
            self.reset_gui_values()
            self.update_streak_values()
            self.create_time_spent_graph()
            self.create_weekday_graph()
            
            self.notification_limit_on = False
        else:
            print("No data to save. (time less than 1m)")

    
    def send_notification(self, title, message):
        toast = Notification(app_id=self.APPNAME, title=title, msg=message)
        toast.show()
        self.notification_limit_on = True
        print("Notification " + title + " sent.")


    def reset_data(self):
        del self.workbook[self.workbook.active.title]
        self.workbook.create_sheet()
        self.worksheet = self.workbook.active

        self.data_manager.reset_data(self.workbook, self.worksheet)

        self.reset_gui_values()
        self.create_time_spent_graph()
        self.create_weekday_graph()


    #Get all buttons other than tab buttons
    def create_widget_list(self):
        frame_list = []
        for frame in self.WINDOW.winfo_children():
            if isinstance(frame, ctk.CTkFrame):
                frame_list.append(frame)
        frame_list.pop(0)
        
        widgets = []

        def _get_widgets(frame):
            for child in frame.winfo_children():
                widgets.append(child)
                if isinstance(child, ctk.CTkFrame):
                    _get_widgets(child)

        for frame in frame_list:
            _get_widgets(frame)

        for widget in widgets:
            if isinstance(widget, ctk.CTkButton):
                self.data_manager.widget_list.append(widget)


    def save_on_quit(self):
        self.save_data()
        print("Data saved on exit.")

        self.workbook.save(self.data_file)
        self.WINDOW.destroy()


    def run(self):
        self.WINDOW.mainloop()


if __name__ == "__main__":
    App().run()
